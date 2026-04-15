"""
=======================================================
  日本株 中長期銘柄選定ツール - リアルタイム版 v3
  ① Yahoo!ニュース感情AIスコア（日英対応）
  ② 景気サイクル連動セクターローテーション
  ③ 日足テクニカル分析（移動平均・ダウ理論・RSI・MACD）
  ④ 出来高急増検知（スマートマネー動向）
  ⑤ EPS成長率トレンド（3期分析）
  ⑥ 前回結果との差分ハイライト（順位変動）
  ⑦ セクター別サマリーシート
  ⑧ 価格モメンタム評価（1M/3M/6M/12M）
  ⑨ Google Trends SNS注目度スコア
  ⑩ スコア根拠コメント自動生成（強み・弱みを日本語で要約） [NEW]
=======================================================

【実行方法】
  1. pip install yfinance openpyxl tqdm pytrends
  2. python japan_stock_realtime.py           # 対話でフェーズ選択
     python japan_stock_realtime.py 拡張期    # フェーズを直接指定
  3. 同フォルダに japan_stock_YYYYMMDD_HHMMSS.xlsx が生成されます
  4. tickers.csv を編集して銘柄を自由に追加・削除できます

【所要時間】約8〜15分（SNSスコア取得込み）
【スコア上限】最大140点（旧120点から拡張）
"""

import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time, os, sys, csv, json
from datetime import datetime
from collections import defaultdict

# ─── オプションライブラリ ──────────────────────────────────────────────────────
try:
    from tqdm import tqdm
    USE_TQDM = True
except ImportError:
    USE_TQDM = False

try:
    from pytrends.request import TrendReq
    USE_PYTRENDS = True
except ImportError:
    USE_PYTRENDS = False

# ─── 設定 ─────────────────────────────────────────────────────────────────────
ENABLE_SNS_SCORE = True   # False にすると SNS スコア取得をスキップ（高速化）

# ─── パス定数 ─────────────────────────────────────────────────────────────────
_DIR           = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE    = os.path.join(_DIR, f"japan_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
TICKERS_CSV    = os.path.join(_DIR, "tickers.csv")
PREV_RANKS_FILE = os.path.join(_DIR, "prev_ranks.json")

# ─── デフォルト銘柄リスト ─────────────────────────────────────────────────────
DEFAULT_TICKER_LIST = [
    ("7203.T","トヨタ自動車","一般消費財"),
    ("6758.T","ソニーグループ","一般消費財"),
    ("9984.T","ソフトバンクグループ","通信"),
    ("8306.T","三菱UFJフィナンシャルG","金融"),
    ("4063.T","信越化学工業","素材"),
    ("6861.T","キーエンス","情報技術"),
    ("9433.T","KDDI","通信"),
    ("7974.T","任天堂","一般消費財"),
    ("6367.T","ダイキン工業","資本財"),
    ("4519.T","中外製薬","ヘルスケア"),
    ("8035.T","東京エレクトロン","情報技術"),
    ("6098.T","リクルートHD","情報技術"),
    ("4661.T","オリエンタルランド","一般消費財"),
    ("8058.T","三菱商事","資本財"),
    ("6954.T","ファナック","資本財"),
    ("4502.T","武田薬品工業","ヘルスケア"),
    ("7267.T","本田技研工業","一般消費財"),
    ("6501.T","日立製作所","資本財"),
    ("9022.T","JR東海","資本財"),
    ("3382.T","セブン&アイHD","生活必需品"),
    ("2914.T","JT","生活必需品"),
    ("9432.T","NTT","通信"),
    ("6702.T","富士通","情報技術"),
    ("8411.T","みずほFG","金融"),
    ("4543.T","テルモ","ヘルスケア"),
    ("8766.T","東京海上HD","金融"),
    ("5108.T","ブリヂストン","素材"),
    ("7751.T","キヤノン","資本財"),
    ("4452.T","花王","生活必需品"),
    ("8802.T","三菱地所","不動産"),
    ("6503.T","三菱電機","資本財"),
    ("4901.T","富士フイルムHD","情報技術"),
    ("2802.T","味の素","生活必需品"),
    ("7733.T","オリンパス","ヘルスケア"),
    ("6326.T","クボタ","資本財"),
    ("8031.T","三井物産","資本財"),
    ("4568.T","第一三共","ヘルスケア"),
    ("3659.T","ネクソン","情報技術"),
    ("6301.T","コマツ","資本財"),
    ("4911.T","資生堂","生活必需品"),
    ("8001.T","伊藤忠商事","資本財"),
    ("5401.T","日本製鉄","素材"),
    ("9101.T","日本郵船","資本財"),
    ("6594.T","ニデック","資本財"),
    ("7741.T","HOYA","ヘルスケア"),
    ("4523.T","エーザイ","ヘルスケア"),
    ("6869.T","シスメックス","ヘルスケア"),
    ("6857.T","アドバンテスト","情報技術"),
    ("9020.T","JR東日本","資本財"),
    ("6752.T","パナソニックHD","一般消費財"),
]

# ─── セクターローテーション ───────────────────────────────────────────────────
SECTOR_ROTATION = {
    "回復期": {"◎":["素材","エネルギー","金融"],     "○":["資本財","情報技術"],    "△":["一般消費財","不動産"],  "×":["生活必需品","公益","通信","ヘルスケア"]},
    "拡張期": {"◎":["情報技術","一般消費財","資本財"],"○":["金融","エネルギー"],   "△":["素材","不動産"],        "×":["生活必需品","公益","通信","ヘルスケア"]},
    "後退期": {"◎":["生活必需品","ヘルスケア","公益"],"○":["通信","不動産"],       "△":["金融","素材"],          "×":["情報技術","一般消費財","資本財","エネルギー"]},
    "不況期": {"◎":["公益","通信","生活必需品"],      "○":["ヘルスケア"],           "△":["金融","不動産"],        "×":["情報技術","一般消費財","資本財","素材","エネルギー"]},
}
SECTOR_SCORE_MAP = {"◎": 20, "○": 14, "△": 7, "×": 0}
VALID_PHASES    = ["回復期", "拡張期", "後退期", "不況期"]

# ─── ニュース感情ワード ────────────────────────────────────────────────────────
POS_WORDS_JA = ["増益","最高益","上方修正","増配","新製品","提携","買収","受注","拡大","成長","回復","黒字","好調","増収","上昇","強化","刷新","供給拡大","新事業","投資拡大","過去最高","躍進","効率化","DX","増額"]
NEG_WORDS_JA = ["減益","下方修正","赤字","訴訟","リコール","損失","不祥事","減配","縮小","悪化","後退","不振","削減","人員削減","撤退","苦戦","低迷","懸念","損害","下落","赤字転落","業績悪化","コスト増","リストラ","格下げ"]
POS_WORDS_EN = ["profit","record","beat","upgrade","raise","increase","growth","expand","surge","rally","dividend","buyback","acquisition","launch","strong","outperform","raised guidance","raised forecast","exceed","positive","recovery","upside","momentum","revised up","higher","gain","boom"]
NEG_WORDS_EN = ["loss","decline","miss","downgrade","cut","reduce","layoff","recall","lawsuit","penalty","warning","risk","weak","disappoint","below","revised down","lower guidance","deficit","scandal","fraud","concern","slump","drop","fell","plunge","restructure","write-off","impairment"]


# ════════════════════════════════════════════════════════════════════════════════
#  CSV銘柄管理
# ════════════════════════════════════════════════════════════════════════════════
def load_tickers():
    if os.path.exists(TICKERS_CSV):
        tickers = []
        with open(TICKERS_CSV, encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                if len(row) >= 3 and row[0].strip() and not row[0].strip().startswith("#"):
                    tickers.append((row[0].strip(), row[1].strip(), row[2].strip()))
        if tickers:
            print(f"  tickers.csv から {len(tickers)} 銘柄を読み込みました。")
            return tickers
    with open(TICKERS_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["# コード", "銘柄名", "セクター  ← この行はコメント。以下を編集して銘柄追加・削除可"])
        w.writerows(DEFAULT_TICKER_LIST)
    print(f"  tickers.csv を生成しました（銘柄の追加・削除はこのファイルを編集してください）。")
    return DEFAULT_TICKER_LIST


# ════════════════════════════════════════════════════════════════════════════════
#  差分管理
# ════════════════════════════════════════════════════════════════════════════════
def load_prev_ranks():
    if os.path.exists(PREV_RANKS_FILE):
        try:
            with open(PREV_RANKS_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_prev_ranks(scored_data):
    # scored_data tuple: (total,sv,sg,sp,sn,ss,st,sm,ssns,grade,d,ns,nt,tech,rank_chg)
    # d は index 10
    ranks = {item[10]["code"]: rank + 1 for rank, item in enumerate(scored_data)}
    with open(PREV_RANKS_FILE, "w", encoding="utf-8") as f:
        json.dump(ranks, f, ensure_ascii=False)


# ════════════════════════════════════════════════════════════════════════════════
#  フェーズ選択
# ════════════════════════════════════════════════════════════════════════════════
def select_phase():
    if len(sys.argv) > 1:
        phase = sys.argv[1].strip()
        if phase in VALID_PHASES:
            return phase
        print(f"  WARNING: 不明なフェーズ '{phase}'。有効値: {' / '.join(VALID_PHASES)}")

    print("\n  景気フェーズを選択してください:")
    for i, p in enumerate(VALID_PHASES, 1):
        print(f"    {i}. {p}")
    print(f"  番号または名前を入力（デフォルト: 拡張期 [Enter]）: ", end="", flush=True)
    try:
        ans = input().strip()
        if not ans:            return "拡張期"
        if ans.isdigit() and 1 <= int(ans) <= len(VALID_PHASES):
            return VALID_PHASES[int(ans) - 1]
        if ans in VALID_PHASES:
            return ans
    except (EOFError, KeyboardInterrupt):
        pass
    return "拡張期"


# ════════════════════════════════════════════════════════════════════════════════
#  データ取得
# ════════════════════════════════════════════════════════════════════════════════
def get_sector_grade(sector, phase):
    for grade, sectors in SECTOR_ROTATION[phase].items():
        if sector in sectors:
            return grade
    return "△"


def fetch_news_for_ticker(ticker_code):
    try:
        tk = yf.Ticker(ticker_code)
        news_list = tk.news
        if not news_list:
            return 0, ["ニュースなし"]
        titles, pos_count, neg_count = [], 0, 0
        cutoff = time.time() - 60 * 86400
        for article in news_list[:30]:
            if isinstance(article, dict):
                content  = article.get("content", article)
                pub_time = content.get("pubDate", "") or article.get("providerPublishTime", 0)
                if isinstance(pub_time, (int, float)) and pub_time > 0 and pub_time < cutoff:
                    continue
                title = content.get("title", "") or article.get("title", "")
            else:
                continue
            if not title:
                continue
            titles.append(title)
            tl = title.lower()
            for w in POS_WORDS_JA:
                if w in title: pos_count += 1
            for w in NEG_WORDS_JA:
                if w in title: neg_count += 1
            for w in POS_WORDS_EN:
                if w in tl: pos_count += 1
            for w in NEG_WORDS_EN:
                if w in tl: neg_count += 1
        total = pos_count + neg_count
        score = 0 if total == 0 else max(-100, min(100, int(((pos_count - neg_count) / total) * 100)))
        return score, titles[:5] if titles else ["取得不可"]
    except Exception as e:
        return 0, [f"エラー:{str(e)[:40]}"]


def fetch_stock_data(ticker_code, name, sector):
    try:
        tk   = yf.Ticker(ticker_code)
        info = tk.info
        per        = info.get("trailingPE") or info.get("forwardPE") or 0
        pbr        = info.get("priceToBook") or 0
        roe        = (info.get("returnOnEquity") or 0) * 100
        rev_growth = (info.get("revenueGrowth") or 0) * 100
        op_margin  = (info.get("operatingMargins") or 0) * 100
        div_yield  = (info.get("dividendYield") or 0) * 100
        eps_growth = (info.get("earningsGrowth") or info.get("earningsQuarterlyGrowth") or 0) * 100
        current_price   = info.get("currentPrice") or info.get("regularMarketPrice") or 0
        prev_close      = info.get("previousClose") or 0
        price_change    = round(current_price - prev_close, 1) if current_price and prev_close else 0
        price_change_pct = round((price_change / prev_close) * 100, 2) if prev_close else 0
        return {
            "code": ticker_code, "name": name, "sector": sector,
            "per": round(per, 1), "pbr": round(pbr, 1), "roe": round(roe, 1),
            "rev_growth": round(rev_growth, 1), "op_margin": round(op_margin, 1),
            "div_yield": round(div_yield, 1), "eps_growth": round(eps_growth, 1),
            "current_price": current_price, "prev_close": prev_close,
            "price_change": price_change, "price_change_pct": price_change_pct,
            "week52_high": info.get("fiftyTwoWeekHigh") or 0,
            "week52_low":  info.get("fiftyTwoWeekLow")  or 0,
        }
    except Exception as e:
        print(f"  ERR {ticker_code}: {e}")
        return None


def fetch_technical_data(ticker_code):
    """テクニカル分析 ＋ モメンタム評価（1M/3M/6M/12Mリターン）[⑧]"""
    try:
        tk = yf.Ticker(ticker_code)
        df = tk.history(period="1y", interval="1d")
        if df.empty or len(df) < 30:
            return None
        close = df["Close"]
        cur   = close.iloc[-1]

        # ── 移動平均 ──────────────────────────────────────────────────────────
        ma5   = close.rolling(5).mean().iloc[-1]   if len(df) >= 5   else None
        ma25  = close.rolling(25).mean().iloc[-1]  if len(df) >= 25  else None
        ma75  = close.rolling(75).mean().iloc[-1]  if len(df) >= 75  else None
        ma200 = close.rolling(200).mean().iloc[-1] if len(df) >= 200 else None

        if   ma200 and ma5 and ma25 and ma75 and cur > ma5 > ma25 > ma75 > ma200: ma_order = "完全上昇配列 ◎"
        elif ma75  and ma25 and cur > ma25 > ma75:                                 ma_order = "上昇配列 ○"
        elif ma75  and ma25 and cur < ma25 < ma75:                                 ma_order = "下降配列 ×"
        else:                                                                       ma_order = "混在 △"

        # ── ゴールデン/デッドクロス ───────────────────────────────────────────
        cross_signal = "なし"
        if ma25 and ma75 and len(df) >= 75:
            ma25s = close.rolling(25).mean()
            ma75s = close.rolling(75).mean()
            for i in range(-1, -31, -1):
                try:
                    if   ma25s.iloc[i] > ma75s.iloc[i] and ma25s.iloc[i-1] <= ma75s.iloc[i-1]:
                        cross_signal = f"GC({abs(i)}日前)🟢"; break
                    elif ma25s.iloc[i] < ma75s.iloc[i] and ma25s.iloc[i-1] >= ma75s.iloc[i-1]:
                        cross_signal = f"DC({abs(i)}日前)🔴"; break
                except Exception:
                    break

        # ── RSI ──────────────────────────────────────────────────────────────
        delta = close.diff()
        gain  = delta.clip(lower=0).rolling(14).mean().iloc[-1]
        loss  = (-delta).clip(lower=0).rolling(14).mean().iloc[-1]
        rsi   = round(100 - 100 / (1 + gain / loss), 1) if loss and loss > 0 else 100.0

        # ── MACD ─────────────────────────────────────────────────────────────
        ema12  = close.ewm(span=12, adjust=False).mean()
        ema26  = close.ewm(span=26, adjust=False).mean()
        macd_l = ema12 - ema26
        sig_l  = macd_l.ewm(span=9, adjust=False).mean()
        macd_v = round(macd_l.iloc[-1], 1)
        sig_v  = round(sig_l.iloc[-1], 1)
        hist_v = round(macd_v - sig_v, 1)

        # ── ダウ理論 ─────────────────────────────────────────────────────────
        recent = close.tail(60); pv = 5; peaks = []; troughs = []
        for i in range(pv, len(recent) - pv):
            win = recent.iloc[i - pv: i + pv + 1]
            if recent.iloc[i] == win.max(): peaks.append(recent.iloc[i])
            if recent.iloc[i] == win.min(): troughs.append(recent.iloc[i])
        dow_trend = "判定中 ─"
        if len(peaks) >= 2 and len(troughs) >= 2:
            hh = peaks[-1] > peaks[-2];   hl = troughs[-1] > troughs[-2]
            lh = peaks[-1] < peaks[-2];   ll = troughs[-1] < troughs[-2]
            if   hh and hl:       dow_trend = "上昇トレンド ▲▲"
            elif lh and ll:       dow_trend = "下降トレンド ▼▼"
            elif hh and not hl:   dow_trend = "上昇転換の兆し ▲"
            elif lh and not ll:   dow_trend = "下降転換の兆し ▼"
            else:                 dow_trend = "レンジ相場 ─"

        # ── 出来高急増検知 ────────────────────────────────────────────────────
        volume    = df["Volume"]
        vol_avg90 = volume.tail(90).mean()
        vol_last  = volume.iloc[-1]
        vol_ratio = round(vol_last / vol_avg90, 2) if vol_avg90 > 0 else 1.0
        if   vol_ratio >= 2.0: vol_signal = f"{vol_ratio:.1f}x 🔥急増"
        elif vol_ratio >= 1.5: vol_signal = f"{vol_ratio:.1f}x ↑増加"
        elif vol_ratio >= 0.7: vol_signal = f"{vol_ratio:.1f}x ─普通"
        else:                  vol_signal = f"{vol_ratio:.1f}x ↓低調"

        # ── モメンタム評価（1M/3M/6M/12M） [⑧] ──────────────────────────────
        def period_return(n):
            idx = min(n, len(close) - 1)
            base = close.iloc[-idx - 1] if idx < len(close) else close.iloc[0]
            return round((cur / base - 1) * 100, 1) if base > 0 else 0.0

        r1m  = period_return(21)   # 約1ヶ月
        r3m  = period_return(63)   # 約3ヶ月
        r6m  = period_return(126)  # 約6ヶ月
        r12m = period_return(252)  # 約12ヶ月

        # モメンタムスコア（最大10pt）
        ms = 0
        # 1ヶ月: 最大2pt
        if   r1m >  10: ms += 2
        elif r1m >   3: ms += 1
        elif r1m < -10: ms -= 1
        # 3ヶ月: 最大3pt
        if   r3m >  20: ms += 3
        elif r3m >   8: ms += 2
        elif r3m >   0: ms += 1
        elif r3m < -15: ms -= 1
        # 6ヶ月: 最大3pt
        if   r6m >  30: ms += 3
        elif r6m >  15: ms += 2
        elif r6m >   0: ms += 1
        elif r6m < -20: ms -= 1
        # 12ヶ月: 最大2pt
        if   r12m > 40: ms += 2
        elif r12m > 20: ms += 1
        elif r12m < -30: ms -= 1

        momentum_score = max(0, min(10, ms))

        # ── テクニカルスコア計算 ──────────────────────────────────────────────
        ts = 0
        if   "完全上昇" in ma_order:   ts += 6
        elif "上昇配列" in ma_order:   ts += 4
        elif "混在"     in ma_order:   ts += 2
        if   "GC"       in cross_signal: ts += 4
        elif "DC"       in cross_signal: ts -= 2
        if   "上昇トレンド"   in dow_trend: ts += 5
        elif "上昇転換"       in dow_trend: ts += 3
        elif "レンジ"         in dow_trend: ts += 1
        if   40 <= rsi <= 65:   ts += 3
        elif 30 <= rsi < 40:    ts += 2
        elif rsi > 80:          ts -= 1
        if macd_v > sig_v and hist_v > 0: ts += 2
        if   vol_ratio >= 2.0: ts += 3
        elif vol_ratio >= 1.5: ts += 2
        elif vol_ratio < 0.7:  ts -= 1

        return {
            "ma5": round(ma5, 0) if ma5 else None,
            "ma25": round(ma25, 0) if ma25 else None,
            "ma75": round(ma75, 0) if ma75 else None,
            "ma200": round(ma200, 0) if ma200 else None,
            "rsi": rsi, "macd": macd_v, "macd_signal": sig_v, "macd_hist": hist_v,
            "cross_signal": cross_signal, "dow_trend": dow_trend, "ma_order": ma_order,
            "vol_ratio": vol_ratio, "vol_signal": vol_signal,
            "r1m": r1m, "r3m": r3m, "r6m": r6m, "r12m": r12m,
            "momentum_score": momentum_score,
            "tech_score": max(0, min(20, ts)),
        }
    except Exception:
        return None


def fetch_sns_score(company_name):
    """
    Google Trends で銘柄の国内検索注目度を取得 [⑨]
    ・スコア 0–100（Googleトレンド指数）→ 0–10pt に正規化
    ・トレンド方向（↑上昇 / ─横ばい / ↓低下）を検出
    ・pytrends 未インストール / 取得失敗時はスキップ
    """
    if not USE_PYTRENDS or not ENABLE_SNS_SCORE:
        return 0.0, "─(スキップ)"
    try:
        time.sleep(1.2)   # レート制限対策
        pytrends = TrendReq(hl="ja-JP", tz=540, timeout=(10, 30),
                            retries=2, backoff_factor=0.5)
        kw = company_name[:20]
        pytrends.build_payload([kw], cat=0, timeframe="today 3-m", geo="JP", gprop="")
        df = pytrends.interest_over_time()
        if df.empty or kw not in df.columns:
            return 0.0, "─"
        scores  = df[kw].astype(float)
        avg     = scores.mean()
        recent  = scores.tail(4).mean()   # 直近4週
        if   recent > avg * 1.15: trend_txt = "↑上昇"
        elif recent < avg * 0.85: trend_txt = "↓低下"
        else:                     trend_txt = "─横ばい"
        sns_score = round(min(10.0, avg / 10.0), 1)
        return sns_score, f"GT:{avg:.0f} {trend_txt}"
    except Exception:
        return 0.0, "─(取得失敗)"


# ════════════════════════════════════════════════════════════════════════════════
#  スコアリング
# ════════════════════════════════════════════════════════════════════════════════
def score_value(per, pbr):
    return round(
        max(0, min(10, (30 - per) / 2) if per and per > 0 else 0) +
        max(0, min(10, (5 - pbr) * 2.5) if pbr and pbr > 0 else 0), 1)

def score_growth(rg, om, eg=0):
    return round(
        min(8, max(0, rg * 0.65) if rg else 0) +
        min(8, max(0, om * 0.28) if om else 0) +
        min(4, max(0, eg * 0.25) if eg else 0), 1)

def score_profitability(roe, om):
    return round(
        min(10, max(0, roe * 0.45) if roe else 0) +
        min(10, max(0, om  * 0.28) if om  else 0), 1)

def score_news(ns):
    return round((ns + 100) / 10, 1)

def compute_total(d, ns, phase, tech, sns_score=0.0):
    sv    = score_value(d["per"], d["pbr"])
    sg    = score_growth(d["rev_growth"], d["op_margin"], d.get("eps_growth", 0))
    sp    = score_profitability(d["roe"], d["op_margin"])
    sn    = score_news(ns)
    grade = get_sector_grade(d["sector"], phase)
    ss    = SECTOR_SCORE_MAP[grade]
    st    = tech["tech_score"]     if tech else 0
    sm    = tech["momentum_score"] if tech else 0   # モメンタムスコア [⑧]
    ssns  = round(sns_score, 1)                     # SNS スコア [⑨]
    total = round(sv + sg + sp + sn + ss + st + sm + ssns, 1)
    return total, sv, sg, sp, sn, ss, st, sm, ssns, grade


# ════════════════════════════════════════════════════════════════════════════════
#  スコア根拠コメント生成
# ════════════════════════════════════════════════════════════════════════════════
def generate_comment(d, sv, sg, sp, sn, ss, st, sm, ssns, grade, tech, total):
    """各スコアの主要因を日本語で簡潔にまとめる（Excel コメント列用）"""
    pos = []   # ポジティブ要因
    neg = []   # ネガティブ要因

    # ── バリュー（PER/PBR）──────────────────────────────────────────────────
    per, pbr = d.get("per", 0), d.get("pbr", 0)
    if sv >= 14:
        pos.append(f"超割安(PER{per:.0f}/PBR{pbr:.1f})" if per and pbr else "超割安")
    elif sv >= 8:
        pos.append("割安")
    elif sv <= 2 and (per > 40 or pbr > 4):
        neg.append(f"割高(PER{per:.0f})" if per > 40 else f"割高(PBR{pbr:.1f})")

    # ── 成長性（売上・利益率・EPS）─────────────────────────────────────────
    eg = d.get("eps_growth", 0)
    if sg >= 12:
        pos.append(f"高成長(EPS{eg:+.0f}%)" if eg and abs(eg) > 5 else "高成長")
    elif sg >= 6:
        pos.append("成長良好")
    if eg <= -15:
        neg.append(f"EPS減益({eg:.0f}%)")
    elif eg >= 30 and sg < 12:
        pos.append(f"EPS急伸({eg:+.0f}%)")

    # ── 収益性（ROE・営業利益率）───────────────────────────────────────────
    roe = d.get("roe", 0)
    if sp >= 12:
        pos.append(f"高収益(ROE{roe:.0f}%)")
    elif sp <= 2:
        neg.append("収益性低")

    # ── ニュース感情 ───────────────────────────────────────────────────────
    if sn >= 16:
        pos.append("ニュース強気📰")
    elif sn <= 7:
        neg.append("ニュース弱気")

    # ── セクターローテーション ─────────────────────────────────────────────
    if grade == "◎":
        pos.append("セクター最適◎")
    elif grade == "○":
        pos.append("セクター良好○")
    elif grade == "×":
        neg.append("セクター逆風×")

    # ── テクニカル ─────────────────────────────────────────────────────────
    if tech:
        ma_order     = tech.get("ma_order", "")
        cross_signal = tech.get("cross_signal", "")
        dow_trend    = tech.get("dow_trend", "")
        rsi          = tech.get("rsi", 50)
        vol_ratio    = tech.get("vol_ratio", 1.0)

        if "完全上昇" in ma_order:
            pos.append("完全上昇配列")
        elif "上昇配列" in ma_order:
            pos.append("上昇配列")
        elif "下降配列" in ma_order:
            neg.append("下降配列")

        if "GC" in cross_signal:
            pos.append(f"GC発生({cross_signal.split('(')[1].rstrip(')')if '(' in cross_signal else '直近'})")
        elif "DC" in cross_signal:
            neg.append("DC発生")

        if "上昇トレンド" in dow_trend:
            pos.append("ダウ上昇▲▲")
        elif "下降トレンド" in dow_trend:
            neg.append("ダウ下降▼▼")

        if rsi >= 75:
            neg.append(f"RSI過熱({rsi:.0f})")
        elif rsi <= 32:
            pos.append(f"RSI売られ過ぎ({rsi:.0f})")

        if vol_ratio >= 2.0:
            pos.append(f"出来高急増🔥({vol_ratio:.1f}x)")
        elif vol_ratio >= 1.5:
            pos.append(f"出来高増({vol_ratio:.1f}x)")

    # ── モメンタム ─────────────────────────────────────────────────────────
    if tech:
        r3m, r6m = tech.get("r3m", 0), tech.get("r6m", 0)
        if sm >= 8:
            pos.append(f"強モメンタム(3M{r3m:+.0f}%/6M{r6m:+.0f}%)")
        elif sm >= 5:
            pos.append(f"モメンタム良(3M{r3m:+.0f}%)")
        elif sm <= 1:
            neg.append(f"モメンタム弱(3M{r3m:+.0f}%)")

    # ── SNS注目度 ──────────────────────────────────────────────────────────
    if ssns >= 6:
        pos.append("SNS注目高📈")
    elif ssns >= 3:
        pos.append("SNS注目中")

    # ── コメント組み立て ───────────────────────────────────────────────────
    parts = []
    if pos:
        parts.append("【強】" + "・".join(pos[:4]))  # 最大4要因
    if neg:
        parts.append("【弱】" + "・".join(neg[:2]))  # 最大2要因
    if not parts:
        parts.append(f"各指標バランス型（総合{total:.0f}pt）")

    return "  ".join(parts)


# ════════════════════════════════════════════════════════════════════════════════
#  Excel スタイル定数
# ════════════════════════════════════════════════════════════════════════════════
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
GOLD_FILL = PatternFill("solid", fgColor="FFF2CC")
GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")
YLW_FILL  = PatternFill("solid", fgColor="FFFF00")
BLU_FILL  = PatternFill("solid", fgColor="DDEEFF")
GRY_FILL  = PatternFill("solid", fgColor="F2F2F2")
RED_FILL  = PatternFill("solid", fgColor="FFE0E0")
ORG_FILL  = PatternFill("solid", fgColor="FFF3CD")
UP_FILL   = PatternFill("solid", fgColor="C8F0C8")
DN_FILL   = PatternFill("solid", fgColor="FFCDD2")
NEW_FILL  = PatternFill("solid", fgColor="B3E5FC")
SNS_FILL  = PatternFill("solid", fgColor="EDE7F6")   # SNS列用（薄紫）
MOM_FILL  = PatternFill("solid", fgColor="E8F5E9")   # モメンタム列用（薄緑）
THIN = Side(style="thin",   color="BFBFBF")
MED  = Side(style="medium", color="4472C4")
THIN_BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BDR  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def hdr(ws, r, c, v, w=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = MED_BDR
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def dat(ws, r, c, v, fmt=None, fill=None, bold=False, align="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = THIN_BDR
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:  cell.number_format = fmt
    if fill: cell.fill = fill
    cell.font = Font(bold=bold, name="Arial", size=9)
    return cell


# ════════════════════════════════════════════════════════════════════════════════
#  Excelシート: 銘柄ランキング（22列 A〜V）
# ════════════════════════════════════════════════════════════════════════════════
# scored_data tuple:
#   (total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech, rank_chg)
#    0      1   2   3   4   5   6   7   8     9      10 11  12  13    14

def build_ranking_sheet(wb, scored_data, phase):
    ws = wb.create_sheet("銘柄ランキング")
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    # タイトル（23列: A〜W）
    ws.merge_cells("A1:W1")
    t = ws["A1"]
    t.value = (f"日本株 中長期銘柄選定ツール v3 - リアルタイムスコアランキング"
               f"  ({datetime.now().strftime('%Y/%m/%d %H:%M')} 取得)  【最大140点】")
    t.font = Font(bold=True, size=13, color="1F3864", name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # フェーズ表示
    ws.merge_cells("A2:C2"); ws["A2"].value = "景気フェーズ(変更可):"
    ws["A2"].font = Font(bold=True, size=10, name="Arial")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["D2"] = phase
    ws["D2"].fill = YLW_FILL
    ws["D2"].font = Font(bold=True, color="0000FF", size=11, name="Arial")
    ws["D2"].alignment = Alignment(horizontal="center"); ws["D2"].border = MED_BDR
    ws.merge_cells("E2:W2")
    ws["E2"].value = ("← 回復期/拡張期/後退期/不況期  ｜  ★TOP3  🔥出来高急増"
                      "  ▲順位UP  ▼DOWN  NEW=新規  GT=Googleトレンド指数(0-100)")
    ws["E2"].font = Font(italic=True, color="595959", size=9, name="Arial")
    ws["E2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ヘッダー（23列）
    headers = [
        ("順位",5), ("変動",7), ("コード",10), ("銘柄名",18), ("セクター",13),
        ("現在株価(円)",11), ("前日比(%)",9), ("52週高値",10), ("52週安値",10),
        ("PER(倍)",8), ("PBR(倍)",8), ("ROE(%)",8),
        ("バリュースコア",10), ("成長性スコア",10), ("収益性スコア",10),
        ("ニューススコア",10), ("セクタースコア",10), ("テクニカルスコア",12),
        ("出来高比率",12), ("モメンタム\nスコア",10), ("SNS注目度\nスコア",11),
        ("総合スコア",10), ("スコア根拠コメント",55),
    ]
    for i, (h, w) in enumerate(headers, 1):
        hdr(ws, 3, i, h, w)
    ws.row_dimensions[3].height = 28

    for rank, item in enumerate(scored_data, 1):
        total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech, rank_chg = item
        row      = rank + 3
        row_fill = GOLD_FILL if rank <= 10 else (GRN_FILL if grade == "◎" else None)
        star     = "★ " if rank <= 3 else ""

        # 順位変動
        if rank_chg is None:
            chg_txt, chg_fill = "NEW", NEW_FILL
        elif rank_chg > 0:
            chg_txt, chg_fill = f"▲{rank_chg}", UP_FILL
        elif rank_chg < 0:
            chg_txt, chg_fill = f"▼{abs(rank_chg)}", DN_FILL
        else:
            chg_txt, chg_fill = "─", row_fill

        dat(ws, row,  1, f"{rank}",               fill=row_fill, bold=rank <= 3)
        dat(ws, row,  2, chg_txt,                 fill=chg_fill, bold=rank_chg != 0)
        dat(ws, row,  3, d["code"],               fill=row_fill)
        dat(ws, row,  4, f"{star}{d['name']}",    fill=row_fill, bold=rank <= 3, align="left")
        dat(ws, row,  5, d["sector"],             fill=row_fill)
        dat(ws, row,  6, d["current_price"] or "-", fmt="#,##0", fill=row_fill, bold=True)
        dat(ws, row,  7, (d["price_change_pct"] / 100) if d["price_change_pct"] else 0,
            fmt="+0.00%;-0.00%;-", fill=row_fill, bold=True)
        dat(ws, row,  8, d["week52_high"] or "-", fmt="#,##0", fill=row_fill)
        dat(ws, row,  9, d["week52_low"]  or "-", fmt="#,##0", fill=row_fill)
        dat(ws, row, 10, d["per"]  or "-",        fmt="0.0",   fill=row_fill)
        dat(ws, row, 11, d["pbr"]  or "-",        fmt="0.0",   fill=row_fill)
        dat(ws, row, 12, (d["roe"] / 100) if d["roe"] else 0, fmt="0.0%", fill=row_fill)
        dat(ws, row, 13, sv,  fmt="0.0", fill=row_fill)
        dat(ws, row, 14, sg,  fmt="0.0", fill=row_fill)
        dat(ws, row, 15, sp,  fmt="0.0", fill=row_fill)
        dat(ws, row, 16, sn,  fmt="0.0", fill=row_fill)
        dat(ws, row, 17, ss,  fmt="0",   fill=row_fill)
        dat(ws, row, 18, st,  fmt="0.0", fill=row_fill)

        # 出来高比率
        vol_sig = tech["vol_signal"] if tech else "-"
        vol_fill = (PatternFill("solid", fgColor="FFD700") if tech and tech["vol_ratio"] >= 2.0
                    else PatternFill("solid", fgColor="DDEEFF") if tech and tech["vol_ratio"] >= 1.5
                    else row_fill)
        dat(ws, row, 19, vol_sig, fill=vol_fill, align="left")

        # モメンタムスコア [⑧]
        mom_detail = ""
        if tech:
            mom_detail = (f"{tech['r1m']:+.0f}%/{tech['r3m']:+.0f}%/"
                          f"{tech['r6m']:+.0f}%/{tech['r12m']:+.0f}%")
        mom_cell_fill = MOM_FILL if sm >= 7 else (DN_FILL if sm <= 2 else row_fill)
        c20 = dat(ws, row, 20, sm, fmt="0.0", fill=mom_cell_fill, bold=sm >= 7)
        if mom_detail:
            c20.comment = None   # コメント追加は省略（openpyxlではcommentオブジェクト必要）

        # SNS注目度スコア [⑨]
        sns_txt = tech["sns_label"] if tech and "sns_label" in tech else (
            f"{ssns:.1f}" if ssns else "─")
        sns_cell_fill = SNS_FILL if ssns >= 5 else row_fill
        dat(ws, row, 21, ssns, fmt="0.0", fill=sns_cell_fill)

        dat(ws, row, 22, total, fmt="0.0", fill=row_fill, bold=True)

        # スコア根拠コメント（col 23）
        comment_txt = generate_comment(d, sv, sg, sp, sn, ss, st, sm, ssns, grade, tech, total)
        comment_fill = GOLD_FILL if rank <= 10 else (GRN_FILL if grade == "◎" else GRY_FILL)
        dat(ws, row, 23, comment_txt, fill=comment_fill, align="left")

        ws.row_dimensions[row].height = 16

    return ws


# ════════════════════════════════════════════════════════════════════════════════
#  Excelシート: セクター別サマリー
# ════════════════════════════════════════════════════════════════════════════════
def build_sector_sheet(wb, scored_data, phase):
    ws = wb.create_sheet("セクター分析")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"セクター別分析  ｜  景気フェーズ: {phase}  ｜  {datetime.now().strftime('%Y/%m/%d %H:%M')} 取得"
    t.font = Font(bold=True, size=12, color="1F3864", name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    hdrs = [("セクター",15),("フェーズ判定",11),("銘柄数",8),("平均スコア",10),
            ("最高スコア銘柄",20),("最高スコア",10),("平均PER",9),("平均ROE(%)",10)]
    for i, (h, w) in enumerate(hdrs, 1):
        hdr(ws, 2, i, h, w)
    ws.row_dimensions[2].height = 20

    sector_data = defaultdict(list)
    for item in scored_data:
        total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech, rank_chg = item
        sector_data[d["sector"]].append({
            "name": d["name"], "total": total, "grade": grade,
            "per": d["per"], "roe": d["roe"],
        })

    grade_order  = {"◎": 0, "○": 1, "△": 2, "×": 3}
    grade_colors = {"◎": "E2EFDA", "○": "DDEEFF", "△": "FFF3CD", "×": "FFE0E0"}

    sectors_sorted = sorted(
        sector_data.items(),
        key=lambda x: (grade_order.get(x[1][0]["grade"], 4),
                       -sum(s["total"] for s in x[1]) / len(x[1]))
    )

    for row_idx, (sector, stocks) in enumerate(sectors_sorted, 3):
        grade     = stocks[0]["grade"]
        avg_score = round(sum(s["total"] for s in stocks) / len(stocks), 1)
        best      = max(stocks, key=lambda x: x["total"])
        pers      = [s["per"] for s in stocks if s["per"] > 0]
        roes      = [s["roe"] for s in stocks if s["roe"] > 0]
        avg_per   = round(sum(pers) / len(pers), 1) if pers else 0
        avg_roe   = round(sum(roes) / len(roes), 1) if roes else 0
        fill      = PatternFill("solid", fgColor=grade_colors.get(grade, "FFFFFF"))

        dat(ws, row_idx, 1, sector,        fill=fill, align="left")
        dat(ws, row_idx, 2, grade,         fill=fill, bold=True)
        dat(ws, row_idx, 3, len(stocks),   fill=fill)
        dat(ws, row_idx, 4, avg_score,     fmt="0.0", fill=fill, bold=True)
        dat(ws, row_idx, 5, best["name"],  fill=fill, align="left")
        dat(ws, row_idx, 6, best["total"], fmt="0.0", fill=fill, bold=True)
        dat(ws, row_idx, 7, avg_per if avg_per else "-",
            fmt="0.0" if avg_per else None, fill=fill)
        dat(ws, row_idx, 8, avg_roe / 100 if avg_roe else 0,
            fmt="0.0%" if avg_roe else None, fill=fill)
        ws.row_dimensions[row_idx].height = 16

    return ws


# ════════════════════════════════════════════════════════════════════════════════
#  メイン
# ════════════════════════════════════════════════════════════════════════════════
def main():
    phase       = select_phase()
    ticker_list = load_tickers()
    prev_ranks  = load_prev_ranks()

    sns_available = USE_PYTRENDS and ENABLE_SNS_SCORE

    print(f"\n{'='*62}")
    print(f"  日本株 中長期銘柄選定ツール v3")
    print(f"  景気フェーズ: {phase}  ｜  対象銘柄数: {len(ticker_list)}")
    print(f"  tqdm進捗バー : {'有効' if USE_TQDM else '無効（pip install tqdm）'}")
    print(f"  SNSスコア   : {'有効（Google Trends）' if sns_available else '無効（pip install pytrends）'}")
    print(f"  差分比較    : {'前回データあり' if prev_ranks else '初回実行'}")
    print(f"  スコア上限  : 最大140点")
    print(f"{'='*62}\n")

    all_data   = []
    sns_cache  = {}   # company_name → (sns_score, sns_label)

    if USE_TQDM:
        iterator = tqdm(enumerate(ticker_list, 1), total=len(ticker_list),
                        desc="銘柄取得", unit="銘柄", ncols=75)
    else:
        iterator = enumerate(ticker_list, 1)

    for i, (code, name, sector) in iterator:
        if not USE_TQDM:
            print(f"[{i:2d}/{len(ticker_list)}] {code} {name}...", end=" ", flush=True)

        d = fetch_stock_data(code, name, sector)
        if d is None:
            if not USE_TQDM: print("skip")
            continue

        ns, nt = fetch_news_for_ticker(code)
        tech   = fetch_technical_data(code)

        # SNSスコア取得 [⑨]（同一銘柄名はキャッシュ）
        if sns_available:
            if name not in sns_cache:
                sns_score, sns_label = fetch_sns_score(name)
                sns_cache[name] = (sns_score, sns_label)
            else:
                sns_score, sns_label = sns_cache[name]
        else:
            sns_score, sns_label = 0.0, "─"

        # tech dict に SNS ラベルを付加（Excel表示用）
        if tech:
            tech["sns_label"] = sns_label

        all_data.append((d, ns, nt, tech, sns_score))

        if USE_TQDM:
            ps = f"¥{d['current_price']:,.0f}" if d['current_price'] else "-"
            iterator.set_postfix({"銘柄": name[:8], "株価": ps, "SNS": sns_label[:6]})
        else:
            ps   = f"¥{d['current_price']:,.0f}({d['price_change_pct']:+.2f}%)" if d['current_price'] else "-"
            ts   = f"Tech:{tech['tech_score']}pt" if tech else "Tech:-"
            ms_v = f"Mom:{tech['momentum_score']}pt" if tech else "Mom:-"
            vs   = tech["vol_signal"] if tech else ""
            print(f"{ps}  {ts}  {ms_v}  {vs}  SNS:{sns_label}  News:{ns:+d}")

        time.sleep(0.5)

    print(f"\n{len(all_data)} 銘柄取得完了。Excel 生成中...")

    # スコアリング & ソート
    scored = []
    for d, ns, nt, tech, sns_score in all_data:
        total, sv, sg, sp, sn, ss, st, sm, ssns, grade = compute_total(d, ns, phase, tech, sns_score)
        scored.append((total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech))
    scored.sort(reverse=True)

    # 差分計算
    final = []
    for rank, item in enumerate(scored, 1):
        total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech = item
        code     = d["code"]
        rank_chg = (prev_ranks[code] - rank) if code in prev_ranks else None
        final.append((total, sv, sg, sp, sn, ss, st, sm, ssns, grade, d, ns, nt, tech, rank_chg))

    # 次回差分用に保存
    save_prev_ranks(final)

    # Excel 出力
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    build_ranking_sheet(wb, final, phase)
    build_sector_sheet(wb, final, phase)
    wb.save(OUTPUT_FILE)

    size_kb = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"\n完成: {OUTPUT_FILE}  ({size_kb} KB)")
    print(f"  シート: 銘柄ランキング / セクター分析")
    print(f"  ⑧ モメンタムスコア（1M/3M/6M/12Mリターン）を追加")
    print(f"  ⑨ SNS注目度スコア（Google Trends 国内検索指数）を追加\n")


if __name__ == "__main__":
    main()
